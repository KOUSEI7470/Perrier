import calendar
from datetime import datetime
from pathlib import Path

import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import jpholiday


BASE_DIR = Path(__file__).resolve().parent
KEY_PATH = BASE_DIR / "serviceAccountKey.json"
OUTPUT_PATH = BASE_DIR / "予約スケジュール.xlsx"

SLOTS = [
    "10:00〜11:00",
    "11:00〜12:00",
    "14:00〜15:00",
    "15:30〜16:30",
    "18:00〜19:00",
]


def init_firebase():
    if not firebase_admin._apps:
        cred = credentials.Certificate(str(KEY_PATH))
        firebase_admin.initialize_app(cred)
    return firestore.client()


def to_datetime(value):
    if value is None:
        return ""
    try:
        return value.strftime("%Y/%m/%d %H:%M")
    except Exception:
        return str(value)


def get_target_year_month():
    now = datetime.now()
    return now.year, now.month


def fetch_firestore_data(db):
    reservations = []
    closed_dates = {}
    closed_slots = {}

    for doc in db.collection("reservations").stream():
        data = doc.to_dict()
        reservations.append(data)

    for doc in db.collection("closedDates").stream():
        data = doc.to_dict()
        if data.get("status") == "active":
            closed_dates[data.get("date")] = data

    for doc in db.collection("closedSlots").stream():
        data = doc.to_dict()
        if data.get("status") == "active":
            key = (data.get("date"), data.get("slot"))
            closed_slots[key] = data

    return reservations, closed_dates, closed_slots


def build_lookup(reservations):
    lookup = {}

    for r in reservations:
        if r.get("status") == "active":
            key = (r.get("date"), r.get("slot"))
            lookup[key] = r

    return lookup


def create_excel(year, month, reservations, closed_dates, closed_slots):
    wb = Workbook()
    ws = wb.active
    ws.title = "予約スケジュール"

    title = f"Perrier 予約スケジュール　{year}年{month}月"
    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color="1F4D3A")
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "予約日",
        "曜日",
        "時間枠",
        "ニックネーム",
        "状態",
        "店側情報",
        "予約ID",
        "確認コード",
        "予約送信日時",
        "キャンセル日時",
    ]

    start_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4D3A")
        cell.alignment = Alignment(horizontal="center")

    thin = Side(style="thin", color="CFCFCF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_reserved = PatternFill("solid", fgColor="DDEFD9")
    fill_closed = PatternFill("solid", fgColor="F4CCCC")
    fill_empty = PatternFill("solid", fgColor="FFFFFF")
    fill_today = PatternFill("solid", fgColor="FFF2CC")

    blue_font = Font(color="1F4E79")
    red_font = Font(color="C00000")
    normal_font = Font(color="000000")

    reservation_lookup = build_lookup(reservations)

    row = start_row + 1
    last_day = calendar.monthrange(year, month)[1]
    today_str = datetime.now().strftime("%Y-%m-%d")

    for day in range(1, last_day + 1):
        date_obj = datetime(year, month, day)
        date_str = date_obj.strftime("%Y-%m-%d")
        date_display = date_obj.strftime("%Y/%m/%d")
        weekday = ["月", "火", "水", "木", "金", "土", "日"][date_obj.weekday()]
        holiday_name = jpholiday.is_holiday_name(date_obj.date())

        for slot_index, slot in enumerate(SLOTS):
            r = reservation_lookup.get((date_str, slot))
            closed_day = closed_dates.get(date_str)
            closed_slot = closed_slots.get((date_str, slot))

            ws.cell(row, 1, date_display if slot_index == 0 else "")
            ws.cell(row, 2, weekday if slot_index == 0 else "")
            ws.cell(row, 3, slot)

            if closed_day:
                status = "予約不可"
                shop_info = closed_day.get("reason", "休業日")
                fill = fill_closed
            elif closed_slot:
                status = "予約不可"
                shop_info = closed_slot.get("reason", "店側都合")
                fill = fill_closed
            elif r:
                status = "予約済み"
                shop_info = ""
                fill = fill_reserved
            else:
                status = ""
                shop_info = ""
                fill = fill_empty

            ws.cell(row, 4, r.get("nickname", "") if r else "")
            ws.cell(row, 5, status)
            ws.cell(row, 6, shop_info)
            ws.cell(row, 7, r.get("reservationId", "") if r else "")
            ws.cell(row, 8, r.get("code", "") if r else "")
            ws.cell(row, 9, to_datetime(r.get("reservationSentAt")) if r else "")
            ws.cell(row, 10, to_datetime(r.get("canceledAt")) if r else "")

            for col in range(1, 11):
                cell = ws.cell(row, col)
                cell.border = border
                cell.fill = fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            if date_str == today_str:
                for col in range(1, 11):
                    ws.cell(row, col).fill = fill_today

            if weekday == "土":
                ws.cell(row, 1).font = blue_font
                ws.cell(row, 2).font = blue_font
            elif weekday == "日" or holiday_name:
                ws.cell(row, 1).font = red_font
                ws.cell(row, 2).font = red_font
            else:
                ws.cell(row, 1).font = normal_font
                ws.cell(row, 2).font = normal_font

            row += 1

    widths = {
        "A": 14,
        "B": 8,
        "C": 16,
        "D": 18,
        "E": 12,
        "F": 22,
        "G": 16,
        "H": 12,
        "I": 20,
        "J": 20,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:J{row - 1}"

    for r in range(4, row):
        ws.row_dimensions[r].height = 24

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return wb


def main():
    if not KEY_PATH.exists():
        raise FileNotFoundError(f"秘密鍵JSONが見つかりません: {KEY_PATH}")

    year, month = get_target_year_month()

    db = init_firebase()
    reservations, closed_dates, closed_slots = fetch_firestore_data(db)

    wb = create_excel(year, month, reservations, closed_dates, closed_slots)
    wb.save(OUTPUT_PATH)

    print(f"Excelを作成しました: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()